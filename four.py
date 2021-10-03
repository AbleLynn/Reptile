# -*- codeing = utf-8 -*-
# @Time : 2021/9/27 19:27
# @Auther : AbleLynn
# @File : four.py
# @software : PyCharm
#导入相关库
import requests
#解析数据内容
from lxml import etree
import json

from matplotlib import pyplot as plt
from openpyxl import Workbook
#数据整理
import pandas as pd
#生成词云
from openpyxl import load_workbook
from wordcloud import WordCloud
#数据分析与可视化
# import pandas as pd
# import matplotlib.pyplot as plt
from matplotlib import rcParams
#绘制国内疫情地图
from pyecharts import options as opts
from pyecharts.charts import Map, Geo
from pyecharts.globals import ThemeType, ChartType  # 主题
from snapshot_selenium import snapshot as driver
import xlrd



"""
1.数据爬取和采集
"""
#爬取的网址（百度疫情）
url="https://voice.baidu.com/act/newpneumonia/newpneumonia/?from=osari_pc_3"
#伪装请求头
"""
为什么要设置headers?
在请求网页爬取的时候，输出的text信息中会出现抱歉，无法访问等字眼，这就是禁止爬取，需要通过反爬机制去解决这个问题。
headers是解决requests请求反爬的方法之一，相当于我们进去这个网页的服务器本身，假装自己本身在爬取数据。
对反爬虫网页，可以设置一些headers信息，模拟成浏览器取访问网站 。
注意：headers中有很多内容，主要常用的就是user-agent 和 host，他们是以键对的形式展现出来，如果user-agent 以字典键对形式作为headers的内容，就可以反爬成功，就不需要其他键对；否则，需要加入headers下的更多键对形式。
"""
headers ={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:92.0) Gecko/20100101 Firefox/92.0'}
#获取网页地址
response=requests.get(url,timeout=30,headers=headers)
# print(response.text)

#json一层一层剥开字典
html = etree.HTML(response.text)
#在网页中寻找我们想要的数据(可以找到对应标签右键复制xpath)
result = html.xpath('//*[@id="captain-config"]/text()') #//*[@id="captain-config"] 注意尾部要加/text()转换为text格式
# print(result)
#result[0]不是真正的字典类型，而是json字符串
#需要通过json.loads转化为python的字典类型
result = json.loads(result[0])
# print(result)
#一层一层找自己所用到的数据
result_out = result["component"][0]
# print(result_out)
#获取国内疫情数据
result_in = result["component"][0]["caseList"]   #国外组caseOutsideList
# print(result_in)

#把数据写到excel表格中
#创建工作簿
wb = Workbook()
#使用工作表
ws = wb.active
#改工作表名称
ws.title = "国内疫情"
#使用append生成表第一行
ws.append(['省份','累计确诊','死亡','治愈','现有确诊','新增','本土新增','境外输入'])
#循环遍历数据，按照位置顺序，将数据加入到excel中
for each in result_in:
    temp_list = [each['area'],each['confirmed'],each['died'],
                 each['crued'],each['curConfirm'],each['confirmedRelative'],
                 each['nativeRelative'],each['overseasInputRelative']]
    ws.append(temp_list)
wb.save("china_data.xlsx")

"""
2.对数据进行清洗
"""
#数据清洗
titanic = pd.DataFrame(pd.read_excel('china_data.xlsx'))
titanic.head()
# print(titanic.head(34))   #.head()默认读取前五行数据 如果要读取多行数据 eg 10行  则为.head(10)
#查看各列数据类型
# titanic.info()
titanic.isnull().head(34)
# print(titanic.isnull().head(34))
#只显示存在缺失值的行列，清楚的确定缺失值的位置
#[clean.isnull().values==True]是条件表达式
titanic[titanic.isnull().values==True].head(34)
titanic.head(34)
# print(titanic.head(34))
#统计各列的空值情况
# print('\n===各列的空值情况如下：===')
titanic.isnull().sum()
# print(titanic.isnull().sum())
#境外输入有六个个空值
# titanic['境外输入'].isnull().value_counts()
# print('\n',titanic['境外输入'].isnull().value_counts())
#使用fillna方法为境外输入字段填充0
titanic['境外输入'] = titanic['境外输入'].fillna('0')
titanic.head(34)
# print(titanic.head(34))
#数据保存
#在完成数据清洗之后，一般会把结果再以csv的格式保存下来，以便后续其他程序的处理
#同样，pandas提供了非常好用的方法
#windows下运行，需要转码，系统默认是gbk，需要手动设置；注意：utf-8还是乱码，得设置为utf-8-sig
titanic.to_csv('china_data_csv.csv',encoding='utf_8_sig')

# """
# 3.生成词云
# """
# wb = load_workbook("china_data.xlsx")
# ws = wb["国内疫情"]
# #应为生成一个词云图，需要一个字典格式{词：词频，词：词频}
# #{省份:累计确诊数}  注：词频是数字类型
# #创建一个空字典
# free={}
# #遍历工作表数据
# for row in ws.values:
# # 不取省份
#     if row[0] != "省份":
#         free[row[0]] = int(row[1])
# # print(free)
# #生成词云
# wc=WordCloud(font_path="msyh.ttc",width=400,height=600,background_color="white")
# wc.generate_from_frequencies(free)
# #保存文件
# wc.to_file("国内疫情数据.png")
# plt.imshow(wc)  #对图像进行处理，并显示其格式
# plt.axis('off')   #隐藏词云图片的坐标轴边界
# plt.show()   #显示处理后的词云图片

"""
4.数据可视化
"""
#正常显示中文
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei']
plt.rcParams['axes.unicode_minus']=False
#绘制各省累计确诊人数垂直柱状图
rcParams['font.family'] = 'simhei'
people = pd.read_excel('china_data.xlsx')
people.plot.bar(x='省份',y='累计确诊')
plt.xticks(rotation=360)
plt.title('各省累计确诊人数')
plt.show()
#绘制各省死亡人数垂直柱状图
rcParams['font.family'] = 'simhei'
people = pd.read_excel('china_data.xlsx')
people.plot.bar(x='省份',y='死亡')
plt.xticks(rotation=360)
plt.title('各省死亡人数')
plt.show()
#绘制各省治愈人数垂直柱状图
rcParams['font.family'] = 'simhei'
people = pd.read_excel('china_data.xlsx')
people.plot.bar(x='省份',y='治愈')
plt.xticks(rotation=360)
plt.title('各省治愈人数')
plt.show()
#绘制各省现有确诊人数垂直柱状图
rcParams['font.family'] = 'simhei'
people = pd.read_excel('china_data.xlsx')
people.plot.bar(x='省份',y='现有确诊')
plt.xticks(rotation=360)
plt.title('各省现有确诊人数')
plt.show()
#绘制各省新增确诊人数垂直柱状图
rcParams['font.family'] = 'simhei'
people = pd.read_excel('china_data.xlsx')
people.plot.bar(x='省份',y='新增')
plt.xticks(rotation=360)
plt.title('各省新增确诊人数')
plt.show()
#绘制累计确诊人数，现有确诊人数，死亡人数，治愈人数的饼图
a = [result_in[0]['confirmed'],result_in[10]['curConfirm'],
     result_in[1]['died'],result_in[2]['crued'],]
plt.pie(a,labels=['累计确诊人数','现有确诊人数','死亡人数','治愈人数'])
plt.show()

#绘制国内疫情地图
filename = 'china_data.xlsx'
file = xlrd.open_workbook(filename)
sheet = file.sheet_by_name('国内疫情')
cityname = sheet.col_values(0)  #获取省份名
# print(cityname)
number = sheet.col_values(1)  # 获取省份累计确诊人数
# print(number)
data = []
for i in range(1,len(cityname)):
    list = []
    list.append(cityname[i])
    list.append(number[i])
    data.append(list)
    # print(list)
#设置地图参数
map = (
    Map(init_opts=opts.InitOpts(bg_color="#FFFAFA", theme=ThemeType.ESSOS, width="1000"))
    .add("累计确诊",data)
    .set_global_opts(
    title_opts=opts.TitleOpts(title=filename[0:5] + "国内数据的疫情图"),
    visualmap_opts=opts.VisualMapOpts(
        is_piecewise=True,  # 设置是否为分段显示
        # 自定义的每一段的范围，以及每一段的文字，以及每一段的特别的样式。例如：
        pieces=[
                {"min": 20000, "label": '>20000人', "color": "#eb2f06"},
                {"min": 10000, "max": 19999, "label": '10000-20000人', "color": "#FF3030"},  # 不指定 max，表示 max 为无限大（Infinity）。
                {"min": 2000, "max": 9999, "label": '2000-10000人', "color": "#FF4500"},
                {"min": 500, "max": 1999, "label": '500-2000人', "color": "#FF7F50"},
                {"min": 50, "max": 499, "label": '50-500人', "color": "#FFA500"},
                {"min": 1, "max": 49, "label": '1-50人', "color": "#FFDEAD"},
        ],
        # 两端的文本，如['High', 'Low']。
        range_text=['高', '低'],
    ),
    )
)
map.render(r'国内疫情.html')
